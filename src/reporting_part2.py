"""
reporting_part2.py — Part II: Performance statistics, carbon plots and tables.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage


# ---------------------------------------------------------------------------
# Performance statistics
# ---------------------------------------------------------------------------

def prepare_risk_free_rate(risk_free_raw):
    """
    Prepare the monthly risk-free rate series.

    The project file stores dates as YYYYMM and RF values in percent.
    The returned series is indexed by month-end Timestamp and expressed
    as a decimal monthly return.
    """
    rf = risk_free_raw.copy()

    date_col = rf.columns[0]
    value_col = "RF" if "RF" in rf.columns else rf.columns[-1]

    dates = pd.to_datetime(rf[date_col].astype(str), format="%Y%m")
    dates = dates + pd.offsets.MonthEnd(0)

    values = pd.to_numeric(rf[value_col], errors="coerce") / 100.0
    return pd.Series(values.values, index=dates, name="RF").sort_index()


def annualized_stats(returns_series, risk_free_rate=None, periods_per_year=12):
    """
    Compute annualized return, volatility, Sharpe ratio, min and max
    from a monthly return series.
    """
    clean = returns_series.dropna()
    mu = clean.mean() * periods_per_year
    sigma = clean.std() * np.sqrt(periods_per_year)

    if risk_free_rate is None:
        excess = clean
    else:
        clean = clean.copy()
        clean.index = pd.to_datetime(clean.index).to_period("M").to_timestamp("M")
        rf = risk_free_rate.copy()
        rf.index = pd.to_datetime(rf.index).to_period("M").to_timestamp("M")
        rf = rf.reindex(clean.index).ffill()
        excess = clean - rf

    excess_mu = excess.dropna().mean() * periods_per_year
    sr = excess_mu / sigma if sigma > 0 else np.nan
    mn = clean.min()
    mx = clean.max()
    return {"Ann. Return": mu, "Ann. Volatility": sigma, "Sharpe Ratio": sr,
            "Min Monthly": mn, "Max Monthly": mx}


def build_performance_table(returns_dict, risk_free_rate=None):
    """
    Build a summary statistics table for multiple portfolios.

    Parameters
    ----------
    returns_dict : dict {label -> pd.Series of monthly returns}

    Returns
    -------
    pd.DataFrame
    """
    rows = []
    for label, ret in returns_dict.items():
        stats = annualized_stats(ret, risk_free_rate=risk_free_rate)
        stats["Portfolio"] = label
        rows.append(stats)
    df = pd.DataFrame(rows).set_index("Portfolio")
    df = df[["Ann. Return", "Ann. Volatility", "Sharpe Ratio", "Min Monthly", "Max Monthly"]]
    return df


def compute_cumulative_returns(returns_dict, base=1.0):
    """
    Compute cumulative return series (starting at `base`) for each portfolio.

    Parameters
    ----------
    returns_dict : dict {label -> pd.Series of monthly returns}

    Returns
    -------
    dict {label -> pd.Series}
    """
    cumulative = {}
    for label, ret in returns_dict.items():
        cumulative[label] = base * (1 + ret.dropna()).cumprod()
    return cumulative


# ---------------------------------------------------------------------------
# Carbon metrics table
# ---------------------------------------------------------------------------

def build_carbon_metrics_table(waci_dict, cf_dict):
    """
    Build a summary table of WACI and CF averages over the sample.

    Parameters
    ----------
    waci_dict : dict {label -> pd.Series}
    cf_dict   : dict {label -> pd.Series}

    Returns
    -------
    pd.DataFrame
    """
    rows = []
    for label in waci_dict:
        rows.append({
            "Portfolio": label,
            "Avg WACI (tCO2/M$ rev)": waci_dict[label].mean(),
            "Avg CF (tCO2/M$ inv)": cf_dict[label].mean(),
            "Min CF": cf_dict[label].min(),
            "Max CF": cf_dict[label].max(),
        })
    return pd.DataFrame(rows).set_index("Portfolio")


# ---------------------------------------------------------------------------
# Plots
# ---------------------------------------------------------------------------

def plot_cumulative_performance(cumulative_dict, figures_dir, filename="cumulative_part2.png",
                                title="Cumulative Portfolio Performance (2014–2025)",
                                show_plot=False):
    """
    Plot cumulative return series for multiple portfolios.
    """
    fig, ax = plt.subplots(figsize=(12, 6))

    styles = ["-", "--", "-.", ":", (0, (3, 1, 1, 1))]
    colors = ["steelblue", "darkorange", "green", "red", "purple"]

    for i, (label, series) in enumerate(cumulative_dict.items()):
        ax.plot(series.index, series.values,
                label=label,
                linestyle=styles[i % len(styles)],
                color=colors[i % len(colors)],
                linewidth=1.8)

    ax.set_title(title, fontsize=13)
    ax.set_xlabel("Date")
    ax.set_ylabel("Cumulative Return (base = 1)")
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    path = figures_dir / filename
    fig.savefig(path, dpi=150)
    if show_plot:
        plt.show()
    plt.close(fig)
    return path


def plot_carbon_metrics(metric_dict, ylabel, title, figures_dir, filename,
                        show_plot=False):
    """
    Plot a carbon metric (WACI or CF) over time for multiple portfolios.

    Parameters
    ----------
    metric_dict : dict {label -> pd.Series indexed by year}
    """
    fig, ax = plt.subplots(figsize=(10, 5))

    styles = ["-o", "--s", "-.^", ":D", "-*"]
    colors = ["steelblue", "darkorange", "green", "red", "purple"]

    for i, (label, series) in enumerate(metric_dict.items()):
        ax.plot(series.index, series.values,
                styles[i % len(styles)],
                label=label,
                color=colors[i % len(colors)],
                linewidth=1.6, markersize=5)

    ax.set_title(title, fontsize=13)
    ax.set_xlabel("Year")
    ax.set_ylabel(ylabel)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    path = figures_dir / filename
    fig.savefig(path, dpi=150)
    if show_plot:
        plt.show()
    plt.close(fig)
    return path


def plot_all_carbon_metrics(waci_dict, cf_dict, figures_dir, show_plot=False):
    """
    Plot WACI and CF side by side.
    """
    p1 = plot_carbon_metrics(
        waci_dict,
        ylabel="WACI (tCO₂ / M$ revenue)",
        title="Weighted Average Carbon Intensity over Time",
        figures_dir=figures_dir,
        filename="waci_timeseries.png",
        show_plot=show_plot,
    )
    p2 = plot_carbon_metrics(
        cf_dict,
        ylabel="CF (tCO₂ / M$ invested)",
        title="Carbon Footprint over Time",
        figures_dir=figures_dir,
        filename="cf_timeseries.png",
        show_plot=show_plot,
    )
    return p1, p2


# ---------------------------------------------------------------------------
# Table image exports
# ---------------------------------------------------------------------------

def _format_table_for_png(df, table_type):
    """
    Return a display-ready copy of a results table.

    CSV files keep the raw decimals for calculations. PNG exports are formatted
    for quick reading in Finder, Overleaf or a slide/report draft.
    """
    display = df.copy()

    if table_type == "performance":
        percent_cols = ["Ann. Return", "Ann. Volatility", "Min Monthly", "Max Monthly"]
        for col in percent_cols:
            if col in display.columns:
                display[col] = display[col].map(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        if "Sharpe Ratio" in display.columns:
            display["Sharpe Ratio"] = display["Sharpe Ratio"].map(
                lambda x: f"{x:.3f}" if pd.notna(x) else ""
            )
        return display

    numeric = display.select_dtypes(include=[np.number]).columns
    for col in numeric:
        display[col] = display[col].map(lambda x: f"{x:.2f}" if pd.notna(x) else "")
    return display


def export_table_png(df, figures_dir, filename, title, table_type="numeric"):
    """
    Export a DataFrame as a PNG table.
    """
    display = _format_table_for_png(df, table_type)
    table_data = display.reset_index()

    n_rows, n_cols = table_data.shape
    fig_width = max(8, min(18, 1.7 * n_cols))
    fig_height = max(2.8, min(12, 0.45 * n_rows + 1.4))

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis("off")
    ax.set_title(title, fontsize=13, fontweight="bold", pad=12)

    table = ax.table(
        cellText=table_data.values,
        colLabels=table_data.columns,
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.35)

    for (row, _), cell in table.get_celld().items():
        cell.set_edgecolor("#D0D7DE")
        if row == 0:
            cell.set_facecolor("#EAECEF")
            cell.set_text_props(weight="bold", color="#24292F")
        else:
            cell.set_facecolor("#FFFFFF" if row % 2 else "#F6F8FA")

    fig.tight_layout()
    path = figures_dir / filename
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path


def export_part2_table_pngs(performance_summary, carbon_summary, waci_by_year,
                            cf_by_year, figures_dir):
    """
    Export the main Part II result tables as PNG files.
    """
    exported = {}
    exported["performance_summary_part2_png"] = export_table_png(
        performance_summary,
        figures_dir,
        filename="performance_summary_part2.png",
        title="Part II Financial Performance",
        table_type="performance",
    )
    exported["carbon_metrics_summary_part2_png"] = export_table_png(
        carbon_summary,
        figures_dir,
        filename="carbon_metrics_summary_part2.png",
        title="Part II Carbon Metrics Summary",
        table_type="numeric",
    )
    exported["waci_by_year_png"] = export_table_png(
        waci_by_year,
        figures_dir,
        filename="waci_by_year_part2.png",
        title="Part II WACI by Year",
        table_type="numeric",
    )
    exported["cf_by_year_png"] = export_table_png(
        cf_by_year,
        figures_dir,
        filename="cf_by_year_part2.png",
        title="Part II Carbon Footprint by Year",
        table_type="numeric",
    )
    return exported


def export_part2_excel_workbook(returns_dict, performance_summary, carbon_summary,
                                waci_by_year, cf_by_year, excel_dir, figures_dir,
                                output_filename="Part_II_results_filled.xlsx"):
    """
    Export a complete Part II workbook for review.

    There is no official Part II Excel template in the assignment. This workbook
    is a clean control file gathering the final Part II tables and figures.
    """
    excel_path = excel_dir / output_filename
    returns_df = pd.DataFrame(returns_dict)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        performance_summary.to_excel(writer, sheet_name="Performance")
        carbon_summary.to_excel(writer, sheet_name="Carbon summary")
        waci_by_year.to_excel(writer, sheet_name="WACI by year")
        cf_by_year.to_excel(writer, sheet_name="CF by year")
        returns_df.to_excel(writer, sheet_name="Monthly returns")

        workbook = writer.book
        summary = workbook.create_sheet("Summary", 0)
        summary["A1"] = "Part II - Portfolio Allocation with Carbon Objective"
        summary["A3"] = "Financial performance"
        summary["A24"] = "Carbon metrics"
        summary["A45"] = "Figures"

        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 24)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    summary = workbook["Summary"]

    figure_positions = [
        ("performance_summary_part2.png", "A4", 720, 230),
        ("carbon_metrics_summary_part2.png", "A25", 720, 230),
        ("cumulative_section4.png", "A46", 620, 310),
        ("cf_section4.png", "A64", 620, 310),
    ]

    for filename, anchor, width, height in figure_positions:
        image_path = figures_dir / filename
        if image_path.exists():
            image = XLImage(str(image_path))
            image.width = width
            image.height = height
            summary.add_image(image, anchor)

    workbook.save(excel_path)
    return excel_path


def export_final_results_workbook(part1_performance, part2_returns,
                                  part2_performance, carbon_summary,
                                  waci_by_year, cf_by_year,
                                  excel_dir, figures_dir,
                                  output_filename="SAAM_Project_GroupN_Final_Results.xlsx"):
    """
    Export one final workbook gathering Part I and Part II results.

    The PNG files remain available separately for the report, but this workbook
    gives one central file for checking all results.
    """
    excel_path = excel_dir / output_filename
    part2_returns_df = pd.DataFrame(part2_returns)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        part1_performance.to_excel(writer, sheet_name="Part I performance", index=False)
        part2_performance.to_excel(writer, sheet_name="Part II performance")
        carbon_summary.to_excel(writer, sheet_name="Part II carbon summary")
        waci_by_year.to_excel(writer, sheet_name="WACI by year")
        cf_by_year.to_excel(writer, sheet_name="CF by year")
        part2_returns_df.to_excel(writer, sheet_name="Monthly returns")

        workbook = writer.book
        summary = workbook.create_sheet("Summary", 0)
        summary["A1"] = "SAAM Project Group N - Final Results"
        summary["A3"] = "Part I"
        summary["A25"] = "Part II - Financial Performance"
        summary["A47"] = "Part II - Carbon Metrics"
        summary["A69"] = "Part II - Time Series Figures"

        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 26)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    summary = workbook["Summary"]

    figure_positions = [
        ("cumulative_portfolio_performance.png", "A4", 620, 310),
        ("performance_summary_part2.png", "A26", 720, 230),
        ("carbon_metrics_summary_part2.png", "A48", 720, 230),
        ("cumulative_section4.png", "A70", 620, 310),
        ("cf_section4.png", "A88", 620, 310),
        ("waci_mv_vw.png", "J4", 560, 280),
        ("cf_mv_vw.png", "J21", 560, 280),
    ]

    for filename, anchor, width, height in figure_positions:
        image_path = figures_dir / filename
        if image_path.exists():
            image = XLImage(str(image_path))
            image.width = width
            image.height = height
            summary.add_image(image, anchor)

    workbook.save(excel_path)
    return excel_path


# ---------------------------------------------------------------------------
# Export utilities
# ---------------------------------------------------------------------------

def export_part2_outputs(returns_dict, waci_dict, cf_dict, tables_dir):
    """
    Export returns, WACI and CF series as CSV files.
    """
    exported = {}

    for label, ret in returns_dict.items():
        safe_label = label.replace(" ", "_").replace("(", "").replace(")", "").replace("/", "")
        path = tables_dir / f"returns_{safe_label}.csv"
        ret.to_csv(path, header=True)
        exported[f"returns_{safe_label}"] = path

    # WACI and CF tables
    waci_df = pd.DataFrame(waci_dict)
    cf_df = pd.DataFrame(cf_dict)

    waci_path = tables_dir / "waci_by_year.csv"
    cf_path = tables_dir / "cf_by_year.csv"
    waci_df.to_csv(waci_path)
    cf_df.to_csv(cf_path)
    exported["waci_by_year"] = waci_path
    exported["cf_by_year"] = cf_path

    return exported
